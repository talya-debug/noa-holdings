# -*- coding: utf-8 -*-
# כתב כמויות הסוללים — בינוי קומה 10
# גרסה פשוטה: טבלה אחת, לירן מתקן ישירות

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

wb = openpyxl.Workbook()

gold = PatternFill('solid', fgColor='FFF2CC')
cat_fill = PatternFill('solid', fgColor='D6EAF8')
green = PatternFill('solid', fgColor='E8F5E9')
header_font = Font(bold=True, size=10)
data_font = Font(size=10)
total_font = Font(bold=True, size=12, color='D4A843')
border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)
center = Alignment(horizontal='center', vertical='center')
rtl = Alignment(horizontal='right', vertical='center', wrap_text=True)

ws = wb.active
ws.title = 'בינוי קומה 10'
ws.sheet_view.rightToLeft = True

# כותרת
ws.merge_cells('A1:H1')
ws['A1'].value = 'כתב כמויות - הסוללים - בינוי קומה 10 | קנ"מ 1:50 | גובה מחיצות 270 סמ'
ws['A1'].font = Font(bold=True, size=13, color='2C3E50')
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws['A1'].fill = gold
ws.row_dimensions[1].height = 30

# הוראה ללירן
ws.merge_cells('A2:H2')
ws['A2'].value = 'לירן: עבור על העמודות. אם נכון — תשאיר. אם לא — תתקן בעמודה הירוקה (H)'
ws['A2'].font = Font(bold=True, size=11, color='FF0000')
ws['A2'].fill = PatternFill('solid', fgColor='FFFF00')
ws.row_dimensions[2].height = 30

# כותרות
headers = ['מס', 'פריט', 'יחידה', 'כמות (נמדד)', 'מחיר יח (דקל 2019)', 'סה"כ', 'איך מדדתי', 'תיקון לירן']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col, value=h)
    cell.font = header_font
    cell.fill = gold
    cell.alignment = center
    cell.border = border

ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 45
ws.column_dimensions['C'].width = 8
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 14
ws.column_dimensions['F'].width = 14
ws.column_dimensions['G'].width = 25
ws.column_dimensions['H'].width = 20

def add_cat_row(row, text):
    ws.merge_cells(f'A{row}:H{row}')
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, size=10, color='2C3E50')
    cell.fill = cat_fill
    cell.alignment = rtl
    cell.border = border

def add_item(row, num, name, unit, qty, price, source):
    ws.cell(row=row, column=1, value=num).font = data_font
    ws.cell(row=row, column=1).alignment = center
    ws.cell(row=row, column=1).border = border

    ws.cell(row=row, column=2, value=name).font = data_font
    ws.cell(row=row, column=2).alignment = rtl
    ws.cell(row=row, column=2).border = border

    ws.cell(row=row, column=3, value=unit).font = data_font
    ws.cell(row=row, column=3).alignment = center
    ws.cell(row=row, column=3).border = border

    c4 = ws.cell(row=row, column=4, value=qty)
    c4.font = Font(bold=True, size=10)
    c4.alignment = center
    c4.border = border

    ws.cell(row=row, column=5, value=price).font = data_font
    ws.cell(row=row, column=5).alignment = center
    ws.cell(row=row, column=5).border = border

    if qty and price:
        ws.cell(row=row, column=6, value=round(qty * price)).font = Font(bold=True, size=10)
    ws.cell(row=row, column=6).border = border

    ws.cell(row=row, column=7, value=source).font = Font(size=9, color='666666')
    ws.cell(row=row, column=7).alignment = rtl
    ws.cell(row=row, column=7).border = border

    # עמודת תיקון — ירוקה, ריקה
    c8 = ws.cell(row=row, column=8, value='')
    c8.fill = green
    c8.border = border

r = 4

# === מחיצות ===
add_cat_row(r, 'מחיצות (מ"ר = אורך שנמדד x גובה 2.7 מ)'); r += 1

items = [
    [1, 'מחיצת אינובייט רצפה-תקרה (אדום בתוכנית)', 'מ"ר', 88, 360, 'PDF: 32.6 מ"א x 2.7'],
    [2, 'קיר גבס דו קרומי 12 סמ + בידוד (כחול בתוכנית)', 'מ"ר', 598, 360, 'PDF: 221.6 מ"א x 2.7'],
    [3, 'קיר גבס ירוק דו קרומי - רטוב (כחול כהה בתוכנית)', 'מ"ר', 253, 460, 'PDF: 93.6 מ"א x 2.7'],
    [4, 'קיר אש 15 סמ (אדום עבה בתוכנית)', 'מ"ר', 72, 600, 'PDF: 26.8 מ"א x 2.7'],
    [5, 'סינר גבס 12 סמ (ירוק בתוכנית)', 'מ"ר', None, 360, 'PDF: 11.8 מ"א, גובה?'],
    [6, 'מחיצה דקורטיבית בלוק זכוכית', 'מ"ר', None, None, 'מקרא, לא נמדד'],
    [7, 'עיטוף קירות גרעין בגבס', 'מ"ר', None, None, 'מהתוכנית'],
]

totals = []
for item in items:
    add_item(r, *item)
    if item[3] and item[4]:
        totals.append(item[3] * item[4])
    r += 1

r += 1

# === דלתות ===
add_cat_row(r, 'דלתות (יחידות — ספירה מהתוכנית)'); r += 1

door_items = [
    [8, 'דלת אינובייט H=270 רוחב 80 סמ', 'יח', 10, 5690, 'ספירה מתוכנית'],
    [9, 'דלת אינובייט H=230 רוחב 80 סמ', 'יח', 6, 5690, 'ספירה מתוכנית'],
    [10, 'דלת חלוקת חללים / מעברים', 'יח', None, 5690, 'מקרא, לא ספרתי'],
    [11, 'דלת אש', 'יח', 1, 4800, 'ספירה מתוכנית'],
]

for item in door_items:
    add_item(r, *item)
    if item[3] and item[4]:
        totals.append(item[3] * item[4])
    r += 1

r += 1

# === פרטים ===
add_cat_row(r, 'פרטים נוספים (מהמקרא)'); r += 1

extra_items = [
    [12, 'שילוט משולב תאורה (SONY)', 'יח', None, None, 'מקרא'],
    [13, 'חיזוקים עבור מסכים', 'יח', None, None, 'מקרא + תוכנית'],
]

for item in extra_items:
    add_item(r, *item)
    r += 1

r += 2

# סיכום
total = sum(totals)
ws.cell(row=r, column=2, value='סה"כ (ממה שנמדד)').font = Font(bold=True, size=12)
ws.cell(row=r, column=6, value=round(total)).font = total_font
ws.cell(row=r, column=2).border = border
ws.cell(row=r, column=6).border = border

# שמירה
output_dir = r'C:\Users\LENOVO\Desktop\cluade\contractor-demo\ללירן'
os.makedirs(output_dir, exist_ok=True)
output = os.path.join(output_dir, 'כתב-כמויות-הסוללים-מתוקן.xlsx')
wb.save(output)
print(f'Saved: {output}')
print(f'Items: {len(items) + len(door_items) + len(extra_items)}')
print(f'Total (measured): {total:,} ILS')
