# -*- coding: utf-8 -*-
# כתב כמויות הסוללים — בינוי קומה 10
# גרסה סופית אחרי למידה + חילוץ מ-PDF
# כמויות נמדדו אוטומטית מ-PDF בקנ"מ 1:50
# גובה מחיצות: 270 סמ (מהתוכנית)
# מחירים: מחירון דקל 2019 פרק 14

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

wb = openpyxl.Workbook()

gold = PatternFill('solid', fgColor='FFF2CC')
cat_fill = PatternFill('solid', fgColor='D6EAF8')
yellow = PatternFill('solid', fgColor='FFFF00')
green_input = PatternFill('solid', fgColor='E8F5E9')
red_fill = PatternFill('solid', fgColor='FFCCCC')
header_font = Font(bold=True, size=10)
data_font = Font(size=10)
total_font = Font(bold=True, size=11, color='D4A843')
border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)
center = Alignment(horizontal='center', vertical='center')
rtl = Alignment(horizontal='right', vertical='center', wrap_text=True)

headers = ['סעיף דקל', 'מס', 'קטגוריה', 'מהות', 'יחידה', 'כמות', 'מחיר יח', 'סה"כ', 'מקור כמות', 'הערות']

def setup_sheet(ws, title):
    ws.sheet_view.rightToLeft = True
    ws.merge_cells('A1:J1')
    ws['A1'].value = title
    ws['A1'].font = Font(bold=True, size=13, color='2C3E50')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].fill = gold
    ws.row_dimensions[1].height = 30
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.fill = gold
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[2].height = 25
    widths = {'A': 14, 'B': 6, 'C': 12, 'D': 45, 'E': 8, 'F': 10, 'G': 12, 'H': 14, 'I': 16, 'J': 30}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def add_cat(ws, row, text):
    ws.merge_cells(f'A{row}:J{row}')
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, size=10, color='2C3E50')
    cell.fill = cat_fill
    cell.alignment = rtl
    cell.border = border

def add_row(ws, row, data, highlight=False):
    for col, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = data_font
        cell.alignment = rtl if col in [4, 9, 10] else center
        cell.border = border
        if highlight:
            cell.fill = yellow
    qty = data[5]
    price = data[6]
    if qty and price:
        ws.cell(row=row, column=8, value=round(qty * price))
        ws.cell(row=row, column=8).font = Font(bold=True, size=10)

def add_summary(ws, row, data_rows):
    total = sum(r[5] * r[6] for r in data_rows if r[5] and r[6])
    ws.cell(row=row, column=4, value='סה"כ (לפני מע"מ)')
    ws.cell(row=row, column=4).font = Font(bold=True, size=12)
    ws.cell(row=row, column=8, value=round(total))
    ws.cell(row=row, column=8).font = total_font
    ws.cell(row=row+1, column=4, value='מע"מ 18%')
    ws.cell(row=row+1, column=8, value=round(total * 0.18))
    ws.cell(row=row+2, column=4, value='סה"כ כולל מע"מ')
    ws.cell(row=row+2, column=4).font = Font(bold=True, size=13)
    ws.cell(row=row+2, column=8, value=round(total * 1.18))
    ws.cell(row=row+2, column=8).font = Font(bold=True, size=13, color='D4A843')

# =============================================
# טאב 1: בינוי קומה 10
# =============================================
ws1 = wb.active
ws1.title = 'בינוי קומה 10'
setup_sheet(ws1, 'כתב כמויות - הסוללים - בינוי קומה 10 | קנ"מ 1:50 | גובה 270 סמ')

# כמויות נמדדו מ-PDF:
# אדום w=1.2 = אינובייט = 32.6 מ"א => 32.6 x 2.7 = 88 מ"ר
# אדום w=6.0 = קיר אש? = 26.8 מ"א => 26.8 x 2.7 = 72 מ"ר
# כחול = גבס = 221.6 מ"א => 221.6 x 2.7 = 598 מ"ר
# כחול כהה = גבס ירוק? = 93.6 מ"א => 93.6 x 2.7 = 253 מ"ר
# ירוק = סינר? = 11.8 מ"א
H = 2.7  # גובה מחיצות במטרים

binuy_data = [
    # --- מחיצות ---
    # אינובייט — נמדד מ-PDF: אדום w=1.2, 32.6 מ"א
    ['14.020.0230', 1, 'מחיצות', 'מחיצת רצפה-תקרה אינובייט, גובה 270 סמ, דו קרומי 12 סמ, כולל בידוד אקוסטי', 'מ"ר',
     round(32.6 * H, 1), 360, None, 'נמדד מ-PDF: 32.6 מ"א אדום x 2.7 גובה',
     'לירן: (1) סעיף 14.020 נכון לאינובייט? (2) 32.6 מ"א נכון?'],

    # קיר גבס — נמדד מ-PDF: כחול, 221.6 מ"א
    ['14.020.0230', 2, 'מחיצות', 'קיר גבס דו קרומי 12 סמ, כולל בידוד אקוסטי 24 קג/מ"ק', 'מ"ר',
     round(221.6 * H, 1), 360, None, 'נמדד מ-PDF: 221.6 מ"א כחול x 2.7 גובה',
     'לירן: כחול = גבס רגיל? 221.6 מ"א נכון?'],

    # קיר גבס ירוק — נמדד מ-PDF: כחול כהה, 93.6 מ"א
    ['14.020.0300', 3, 'מחיצות', 'קיר גבס ירוק דו קרומי (מטבחים/רטוב), כולל בידוד אקוסטי', 'מ"ר',
     round(93.6 * H, 1), 460, None, 'נמדד מ-PDF: 93.6 מ"א כחול כהה x 2.7 גובה',
     'לירן: כחול כהה = גבס ירוק? 93.6 מ"א נכון?'],

    # קיר אש — נמדד מ-PDF: אדום w=6.0, 26.8 מ"א
    ['14.010.0040', 4, 'מחיצות', 'קיר אש 15 סמ', 'מ"ר',
     round(26.8 * H, 1), 600, None, 'נמדד מ-PDF: 26.8 מ"א אדום עבה x 2.7 גובה',
     'לירן: אדום עבה מאוד = קיר אש? 26.8 מ"א נכון?'],

    # סינר — נמדד מ-PDF: ירוק, 11.8 מ"א
    # סינר הוא לא בגובה 270! הוא מתקרה עד 270/230
    ['14.060.0200', 5, 'מחיצות', 'סינר גבס 12 סמ (מתקרה עד H=270/230)', 'מ"ר',
     None, 360, None, 'נמדד מ-PDF: 11.8 מ"א ירוק, גובה סינר לא ידוע',
     'לירן: כמה מ"א סינר? מה גובה כל סינר?'],

    # זכוכית — ממקרא, לא נמדד
    ['', 6, 'מחיצות', 'מחיצה דקורטיבית מבלוק זכוכית', 'מ"ר',
     None, None, None, 'מהמקרא, לא נמדד',
     'לירן: כמה מ"ר? מה סעיף דקל?'],

    # קיר בטון חשוף קיים — לא נמדד
    ['', 7, 'מחיצות', 'קיר בטון חשוף (קיים, שימור)', 'מ"ר',
     None, None, None, 'מהמקרא',
     'לירן: נכנס לכתב כמויות?'],

    # --- דלתות ---
    ['06.012.0010', 8, 'דלתות', 'דלת אינובייט כניסה לחדרים H=270, רוחב 80 סמ', 'יח',
     10, 5690, None, 'ספירה ויזואלית',
     'לירן: 10 נכון? רגיל או נגיש?'],

    ['06.012.0010', 9, 'דלתות', 'דלת אינובייט כניסה לחדרים H=230, רוחב 80 סמ', 'יח',
     6, 5690, None, 'ספירה ויזואלית',
     'לירן: 6 נכון? רגיל או נגיש?'],

    ['06.012.0010', 10, 'דלתות', 'דלת חלוקת חללים/מעברים, רוחב 80 סמ', 'יח',
     None, 5690, None, 'מהמקרא',
     'לירן: כמה?'],

    ['06.033.0094', 11, 'דלתות', 'דלת אש', 'יח',
     1, 4800, None, 'ספירה ויזואלית',
     'לירן: 1 נכון?'],

    # --- פרטים ---
    ['', 12, 'פרטים', 'שילוט משולב תאורה (SONY)', 'יח',
     None, None, None, 'מהמקרא',
     'לירן: כמה? מה המחיר?'],

    ['', 13, 'פרטים', 'חיזוקים עבור מסכים (קיר מחוזק)', 'יח',
     None, None, None, 'מהמקרא+תוכנית',
     'לירן: כמה? סעיף דקל?'],

    ['', 14, 'פרטים', 'עיטוף קירות גרעין בגבס', 'מ"ר',
     None, None, None, 'מהתוכנית',
     'לירן: כמה מ"ר?'],
]

r = 3
add_cat(ws1, r, 'מחיצות — מ"ר (מ"א x גובה 270 סמ) — נמדדו מ-PDF'); r += 1
for row_data in binuy_data[:7]:
    hl = row_data[5] is None  # צהוב = חסר כמות
    add_row(ws1, r, row_data, highlight=hl); r += 1
r += 1
add_cat(ws1, r, 'דלתות — יחידות'); r += 1
for row_data in binuy_data[7:11]:
    hl = row_data[5] is None
    add_row(ws1, r, row_data, highlight=hl); r += 1
r += 1
add_cat(ws1, r, 'פרטים נוספים מהמקרא'); r += 1
for row_data in binuy_data[11:]:
    add_row(ws1, r, row_data, highlight=True); r += 1
r += 2
add_summary(ws1, r, binuy_data)

# =============================================
# טאב 2: שאלות ללירן (רק בינוי)
# =============================================
ws2 = wb.create_sheet('שאלות ללירן')
ws2.sheet_view.rightToLeft = True
ws2.merge_cells('A1:D1')
ws2['A1'].value = 'לירן — מלא את העמודה הירוקה (D) ותחזיר'
ws2['A1'].font = Font(bold=True, size=14, color='FF0000')
ws2['A1'].fill = yellow
ws2.row_dimensions[1].height = 40

q_headers = ['מס', 'נושא', 'שאלה', 'תשובת לירן']
for col, h in enumerate(q_headers, 1):
    cell = ws2.cell(row=2, column=col, value=h)
    cell.font = Font(bold=True, size=11)
    cell.fill = gold
    cell.border = border
    cell.alignment = center

ws2.column_dimensions['A'].width = 6
ws2.column_dimensions['B'].width = 14
ws2.column_dimensions['C'].width = 65
ws2.column_dimensions['D'].width = 40

questions = [
    # אימות צבעים
    ['צבעים', 'אדום (קו דק) = מחיצת אינובייט. נכון?'],
    ['צבעים', 'כחול = קיר גבס דו קרומי 12 סמ. נכון?'],
    ['צבעים', 'כחול כהה = גבס ירוק (רטוב/מטבחים). נכון?'],
    ['צבעים', 'אדום עבה מאוד = קיר אש 15 סמ. נכון?'],
    ['צבעים', 'ירוק = סינר. נכון?'],
    # אימות כמויות
    ['כמויות', 'אינובייט: נמדד 32.6 מ"א מהתוכנית. נכון?'],
    ['כמויות', 'גבס רגיל: נמדד 221.6 מ"א. נכון? (נראה הרבה!)'],
    ['כמויות', 'גבס ירוק: נמדד 93.6 מ"א. נכון?'],
    ['כמויות', 'קיר אש: נמדד 26.8 מ"א. נכון?'],
    ['כמויות', 'סינר: כמה מ"א? מה גובה כל סינר? (גובה סינר ≠ 270!)'],
    ['כמויות', 'בלוק זכוכית: כמה מ"ר?'],
    ['כמויות', 'עיטוף גרעין גבס: כמה מ"ר?'],
    # דלתות
    ['דלתות', 'דלתות H=270: ספרתי 10. נכון?'],
    ['דלתות', 'דלתות H=230: ספרתי 6. נכון?'],
    ['דלתות', 'דלתות חלוקת חללים: כמה?'],
    ['דלתות', 'דלתות אש: 1. נכון?'],
    ['דלתות', 'יש דלתות נגיש? כמה?'],
    # סעיפי דקל
    ['סעיפים', 'אינובייט: סעיף 14.020.0230 (גבס דו קרומי 15 סמ) מתאים? או סעיף אחר?'],
    ['סעיפים', 'קיר אש: סעיף 14.010.0040 מתאים?'],
    ['סעיפים', 'בלוק זכוכית: מה סעיף דקל?'],
    ['סעיפים', 'שילוט SONY: מה המחיר? סעיף?'],
    # כללי
    ['כללי', 'מחירים מדקל 2019. צריך מדד עדכון?'],
]

for i, (topic, q) in enumerate(questions, 1):
    row = i + 2
    ws2.cell(row=row, column=1, value=i).font = data_font
    ws2.cell(row=row, column=1).alignment = center
    ws2.cell(row=row, column=1).border = border
    ws2.cell(row=row, column=2, value=topic).font = data_font
    ws2.cell(row=row, column=2).alignment = center
    ws2.cell(row=row, column=2).border = border
    cell_q = ws2.cell(row=row, column=3, value=q)
    cell_q.font = Font(bold=True, size=10)
    cell_q.fill = yellow
    cell_q.alignment = rtl
    cell_q.border = border
    cell_a = ws2.cell(row=row, column=4, value='')
    cell_a.fill = green_input
    cell_a.border = border

# =============================================
# טאב 3: מה תוקן ולמה
# =============================================
ws3 = wb.create_sheet('מה תוקן')
ws3.sheet_view.rightToLeft = True
ws3.merge_cells('A1:D1')
ws3['A1'].value = 'תיקונים שעשינו — הסבר ללירן'
ws3['A1'].font = Font(bold=True, size=14, color='2C3E50')
ws3['A1'].fill = gold
ws3.row_dimensions[1].height = 35

fix_headers = ['מס', 'מה היה', 'מה עכשיו', 'למה']
for col, h in enumerate(fix_headers, 1):
    cell = ws3.cell(row=2, column=col, value=h)
    cell.font = Font(bold=True, size=11)
    cell.fill = gold
    cell.border = border
    cell.alignment = center

ws3.column_dimensions['A'].width = 6
ws3.column_dimensions['B'].width = 30
ws3.column_dimensions['C'].width = 35
ws3.column_dimensions['D'].width = 45

fixes = [
    ['מחיצות במ"א', 'תוקן למ"ר (מ"א x 2.7 גובה)', 'כתב כמויות = מ"ר. למדנו מבת ים'],
    ['כתבנו "ריינוביט"', 'תוקן ל"אינובייט"', 'קראנו את המקרא מהתוכנית'],
    ['גובה לא ידוע', 'מצאנו 270/280 סמ', 'כתוב בתוכנית: "גובה משתנה 280/270"'],
    ['רק 2 סוגי מחיצות', '7 סוגים מהמקרא', 'חילצנו: אינובייט, גבס, ירוק, אש, סינר, זכוכית, בטון'],
    ['כמויות ידניות (177 מ"א)', 'נמדדו מ-PDF לפי צבע', 'סקריפט אוטומטי: 32.6+221.6+93.6+26.8+11.8'],
    ['שמות דלתות לא מדויקים', 'לפי מקרא: H=270 ו-H=230', 'כלל 3: שמות בדיוק מהמקרא'],
    ['סעיפי 04.020', 'שונה ל-14.020 (גבס)', 'אינובייט = מחיצות גבס, לא בלוקים'],
    ['לא היו פרטים נוספים', 'הוספנו: SONY, חיזוקי מסכים, עיטוף', 'חילצנו מהמקרא של התוכנית'],
]

for i, (was, now, why) in enumerate(fixes, 1):
    row = i + 2
    ws3.cell(row=row, column=1, value=i).font = data_font
    ws3.cell(row=row, column=1).alignment = center
    ws3.cell(row=row, column=1).border = border
    c2 = ws3.cell(row=row, column=2, value=was)
    c2.font = data_font; c2.alignment = rtl; c2.fill = red_fill; c2.border = border
    c3 = ws3.cell(row=row, column=3, value=now)
    c3.font = Font(bold=True, size=10); c3.alignment = rtl; c3.fill = green_input; c3.border = border
    c4 = ws3.cell(row=row, column=4, value=why)
    c4.font = data_font; c4.alignment = rtl; c4.border = border

# שמירה
output_dir = r'C:\Users\LENOVO\Desktop\cluade\contractor-demo\ללירן'
os.makedirs(output_dir, exist_ok=True)
output = os.path.join(output_dir, 'כתב-כמויות-הסוללים-מתוקן.xlsx')
wb.save(output)
print(f'Saved to {output}')
print(f'Tab 1: binuy - {len(binuy_data)} items')
print(f'Tab 2: questions - {len(questions)} questions')
print(f'Tab 3: fixes - {len(fixes)} fixes')
